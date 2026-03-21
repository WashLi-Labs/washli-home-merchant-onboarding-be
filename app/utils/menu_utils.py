import pdfplumber
import openpyxl
import io
import re
from typing import List

# Predefined list of items to look for
MENU_ITEMS_TO_EXTRACT = [
    "Formal Shirt", "Casual Shirt", "Trousers", "Underclothes", 
    "Frock / Dress", "Saree", "Suit (2-Piece)", "Shorts", 
    "Bedsheet", "Denim", "socks", "Jackets", "Towells"
]

def extract_items_from_document(content: bytes, filename: str) -> List[str]:
    """
    Extracts predefined menu items from a PDF or Excel document.
    """
    found_items = []
    
    try:
        if filename.lower().endswith('.pdf') or content.startswith(b'%PDF'):
            found_items = _extract_from_pdf(content)
        elif filename.lower().endswith(('.xlsx', '.xls')):
            found_items = _extract_from_excel(content)
        else:
            # Try both if extension is unknown
            try:
                found_items = _extract_from_pdf(content)
            except:
                try:
                    found_items = _extract_from_excel(content)
                except:
                    pass
    except Exception as e:
        print(f"Error extracting from {filename}: {e}")
        
    return found_items

def _extract_from_pdf(content: bytes) -> List[str]:
    found = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        full_text = ""
        for page in pdf.pages:
            # Try to extract tables first, as it's more structured
            tables = page.extract_tables()
            for table in tables:
                if not table: continue
                # Look for "Item Name" header
                header_row = -1
                item_col = -1
                for r_idx, row in enumerate(table):
                    for c_idx, cell in enumerate(row):
                        if cell and "item name" in str(cell).lower():
                            header_row = r_idx
                            item_col = c_idx
                            break
                    if item_col != -1: break
                
                if item_col != -1:
                    # Extract everything in this column below the header
                    for r_idx in range(header_row + 1, len(table)):
                        item = table[r_idx][item_col]
                        if item and str(item).strip():
                            found.append(str(item).strip())
            
            # If no tables found or no items found in tables, fallback to text search with predefined list
            if not found:
                text = page.extract_text()
                if text:
                    for item in MENU_ITEMS_TO_EXTRACT:
                        if re.search(re.escape(item), text, re.IGNORECASE):
                            if item not in found:
                                found.append(item)
    return list(set(found))

def _extract_from_excel(content: bytes) -> List[str]:
    found = []
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    
    for sheet in wb.worksheets:
        item_col = -1
        header_row = -1
        
        # Search for "Item Name" header in the first 10 rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True)):
            for col_idx, cell in enumerate(row):
                if cell and "item name" in str(cell).lower():
                    item_col = col_idx
                    header_row = row_idx + 1
                    break
            if item_col != -1: break
            
        if item_col != -1:
            # Extract from this column
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                item = row[item_col]
                if item and str(item).strip():
                    found.append(str(item).strip())
        else:
            # Fallback: Search all cells for predefined items if "Item Name" header not found
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        cell_str = str(cell).strip()
                        for item in MENU_ITEMS_TO_EXTRACT:
                            if item.lower() == cell_str.lower():
                                found.append(item)
                                
    return list(set(found))
